# -*- coding: utf-8 -*-
"""
Haengt Dashboard-Neuanlagen (aus lackformeln_eingang) als lesbare Bloecke an die
BEGLEITDATEI Dashboard_Neu.xlsx an. Der Excel-Master wird bewusst NICHT angefasst.

Aufruf:
    python eingang_to_excel.py --in eingang.json --file "Pfad\\Dashboard_Neu.xlsx"

Erwartet eine JSON-Liste mit den Feldern kunde, farbname, system, komponenten,
primer, klarlack, haerter, notiz, erfasst_von, erstellt_am.
Exit-Code 2 = Datei gesperrt (in Excel geoeffnet) -> Aufrufer soll spaeter erneut.
"""
import sys
import os
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

SHEET = "Neuanlagen"


def arg(name, default=None):
    for i, a in enumerate(sys.argv):
        if a == name and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
        if a.startswith(name + "="):
            return a.split("=", 1)[1]
    return default


def load_or_create(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A1"] = "Dashboard-Neuanlagen Lackformeln – bitte in den Excel-Master einsortieren"
    ws["A1"].font = Font(bold=True)
    ws.append([])
    return wb


def main():
    in_path = arg("--in")
    file_path = arg("--file")
    if not in_path or not file_path:
        sys.stderr.write("Aufruf: --in <json> --file <xlsx>\n")
        sys.exit(1)

    with open(in_path, "r", encoding="utf-8") as fh:
        rows = json.load(fh)
    if not rows:
        sys.stderr.write("Keine neuen Eintraege.\n")
        return

    wb = load_or_create(file_path)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.create_sheet(SHEET)

    for r in rows:
        ws.append([r.get("farbname", ""), r.get("kunde", "")])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        if r.get("system"):
            ws.append([r["system"]])
        for k in (r.get("komponenten") or []):
            row = [k.get("code", ""), k.get("gramm", "")]
            if k.get("gruppe"):
                row.append(k["gruppe"])
            ws.append(row)
        for label, key in (("Primer", "primer"), ("Klarlack", "klarlack"),
                           ("Haerter", "haerter"), ("Notiz", "notiz")):
            if r.get(key):
                ws.append([f"{label}: {r[key]}"])
        meta = "Erfasst"
        if r.get("erfasst_von"):
            meta += f" von {r['erfasst_von']}"
        if r.get("erstellt_am"):
            meta += f" am {str(r['erstellt_am'])[:16]}"
        ws.append([meta])
        ws.append([])  # Leerzeile als Trenner

    try:
        wb.save(file_path)
    except PermissionError:
        sys.stderr.write("Dashboard_Neu.xlsx ist geoeffnet/gesperrt – spaeter erneut.\n")
        sys.exit(2)
    sys.stderr.write(f"{len(rows)} Neuanlage(n) angehaengt -> {file_path}\n")


if __name__ == "__main__":
    main()
