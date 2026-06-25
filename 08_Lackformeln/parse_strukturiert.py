# -*- coding: utf-8 -*-
"""
Parser fuer Lackformeln_Strukturiert.xlsx (EINHAUS Oberflaechenveredelung GmbH).

Liest die saubere, strukturierte Eingabedatei (Blatt "Formeln") und erzeugt
Lackformeln als JSON im Schema der Supabase-Tabelle `lackformeln`.

Eingabe-/Parse-Regel:
  - Neuer Farbname in Spalte C  = neue Formel (Kopfzeile mit Kunde/Code/Typ/...).
  - Folgezeilen mit leerem Farbnamen = weitere Komponenten (nur Bereich/Toenpaste/Gramm).
  - geprueft = TRUE, wenn die Spalte "Status" leer ist (orange Formeln haben Status-Text).
  - "Bereich" (Grundton/Effekt) wird als komponenten.gruppe uebernommen.

Aufruf (wie der alte Parser, vom Sync verwendet):
    python parse_strukturiert.py "Pfad\\Lackformeln_Strukturiert.xlsx" --out parsed.json
"""
import sys, os, json, shutil, tempfile
import openpyxl

DEFAULT_XLSX = r"C:\Users\lisina\einhaus-gmbh.de\Lackierung - Dokumente\4. Lackformeln\Lackformeln_Strukturiert.xlsx"
SHEET = "Formeln"
# Spaltenindex (1-basiert)
C_KUNDE,C_CODE,C_NAME,C_TYP,C_SCH,C_SYS,C_BER,C_TP,C_G,C_PRIM,C_KLAR,C_CC,C_HAE,C_VERH,C_VOR,C_NOTIZ,C_STATUS = range(1,18)

def num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    s=str(v).strip().replace(" ","").replace(",", ".")
    try: return float(s)
    except: return None

def txt(v):
    if v is None: return None
    s=str(v).strip()
    return s if s else None

RANK = {"Grundton": 0, "": 1, "Effekt": 2}

def parse_sheet(ws, quelle_blatt):
    records = []; cur = None
    for r in range(2, ws.max_row + 1):
        def cell(c, _r=r): return ws.cell(_r, c).value
        name = txt(cell(C_NAME))
        if name:
            notiz = txt(cell(C_NOTIZ)) or ""
            cc = txt(cell(C_CC)); verh = txt(cell(C_VERH))
            extra = []
            if cc:   extra.append("Klarlack-Toenpaste: %s" % cc)
            if verh: extra.append("Verhaeltnis: %s" % verh)
            if extra:
                notiz = (notiz + " | " if notiz else "") + " | ".join(extra)
            cur = {
                "kunde": txt(cell(C_KUNDE)) or "",
                "farbname": name,
                "farbcode": txt(cell(C_CODE)),
                "system": txt(cell(C_SYS)),
                "typ": txt(cell(C_TYP)),
                "komponenten": [],
                "primer": txt(cell(C_PRIM)),
                "klarlack": txt(cell(C_KLAR)),
                "haerter": txt(cell(C_HAE)),
                "vorlage_nr": txt(cell(C_VOR)),
                "notiz": notiz or None,
                "quelle_blatt": quelle_blatt,
                "quelle_zeile": r,
                "geprueft": txt(cell(C_STATUS)) is None,
            }
            records.append(cur)
        code = txt(cell(C_TP))
        if cur is not None and code:
            cur["komponenten"].append({
                "code": code,
                "gramm": num(cell(C_G)),
                "gruppe": txt(cell(C_BER)) or "",
            })
    records = [x for x in records if x["komponenten"]]
    for x in records:
        x["komponenten"].sort(key=lambda k: RANK.get(k.get("gruppe", ""), 1))
    return records

def main():
    src = sys.argv[1] if len(sys.argv)>1 and not sys.argv[1].startswith("--") else DEFAULT_XLSX
    out = None
    if "--out" in sys.argv:
        out = sys.argv[sys.argv.index("--out")+1]

    # Robust gegen Datei-Sperre (Koloristin hat die Datei evtl. offen): in Temp kopieren.
    tmp = os.path.join(tempfile.gettempdir(), "_lackformeln_struktur_read.xlsx")
    try:
        shutil.copyfile(src, tmp); read_path = tmp
    except Exception:
        read_path = src
    wb = openpyxl.load_workbook(read_path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.stderr.write("Blatt '%s' nicht gefunden\n" % SHEET); sys.exit(1)

    records = parse_sheet(wb[SHEET], "Strukturiert")
    # Archiv-Blatt (Canyon u.a.) ebenfalls einspielen, aber als quelle_blatt="Archiv"
    if "Archiv" in wb.sheetnames:
        archiv = parse_sheet(wb["Archiv"], "Archiv")
        records += archiv
        sys.stderr.write("%d Archiv-Formeln eingelesen\n" % len(archiv))

    data = json.dumps(records, ensure_ascii=False)
    if out:
        with open(out, "w", encoding="utf-8") as f: f.write(data)
    else:
        sys.stdout.write(data)
    aktiv = [x for x in records if x["quelle_blatt"] != "Archiv"]
    ok = sum(1 for x in aktiv if x["geprueft"])
    sys.stderr.write("%d Formeln gesamt (%d aktiv, %d geprueft, %d zu pruefen)\n"
                     % (len(records), len(aktiv), ok, len(aktiv)-ok))

if __name__ == "__main__":
    main()
